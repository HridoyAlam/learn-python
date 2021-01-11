import openpyxl as xl
from openpyxl.chart import BarChart, Reference


#def procces_workbook('transaction.xlsx'):
wb = xl.load_workbook('transaction.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1,1)
    #print(cell.value)

    #how many rows in sheet
    #print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
        #print(row)
    cell = sheet.cell(row,3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected_price
    #print(corrected_price)

values = Reference(sheet,
            min_row = 2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)


chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'g2')

#save as a new file
wb.save('transaction1.xlsx')

#procces_workbook('transaction.xlsx')
